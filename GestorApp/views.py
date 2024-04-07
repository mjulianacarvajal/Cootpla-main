
# gestion de usuarios y accesos en inglés: clases y variables propias de las librerias

import json
from datetime import datetime

from django.contrib import messages
from django.contrib.auth import (
    authenticate, login, logout,
    update_session_auth_hash
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.management import call_command
from django.db.models import Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render



from .forms import (
    ActualizarContrasena, ActualizarPerfil, GuardarEncomienda,
    GuardarProgramacion, GuardarSede, GuardarVehiculo,
    RegistrarUsuario
)
from .models import (
    Conductor, Encomienda, Programacion, Propietario, Sede,
    Vehiculo
)

from openpyxl import *
from openpyxl.styles import *
import getpass




context = {
    'page_title': 'Visor de Viajes Intermunicipales || Cootpla',
}



def login_user(request):
    logout(request)
    resp = {"status":'failed','msg':''}
    username = ''
    password = ''
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                resp['status']='success'
            else:
                resp['msg'] = "Usuario o Contraseña: Incorrecta"
        else:
            resp['msg'] = "Usuario o Contraseña: Incorrecta"
    return HttpResponse(json.dumps(resp),content_type='application/json')


def logoutuser(request: HttpRequest):
    logout(request)
    return redirect('/')

@login_required
def inicio(request: HttpRequest):
    context['page_title'] = 'Inicio'
    #tabla de contenido o dashboard inicio
    context['encomiendas'] = Encomienda.objects.count()
    context['vehiculos'] = Vehiculo.objects.count()
    context['programaciones'] = Programacion.objects.filter(programacion__gt= datetime.today()).count()
    context['username'] = User.objects.count()

    labels = []
    data = []

    queryset = Vehiculo.objects.order_by('estado')
    for veh_c in queryset:
        if veh_c.estado == '1':
            labels.append(veh_c.numero_veh)
            data.append(veh_c.asientos)

    context['labels'] = labels
    context['data'] = data

    return render(request,'core/inicio.html',context)



@login_required
def registrarUsuario(request: HttpRequest):
    user = request.user
    context['page_title'] = "Registrar Usuario"
    if request.method == 'POST':
        data = request.POST.copy()
        form = RegistrarUsuario(data)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            pwd = form.cleaned_data.get('password1')
            loginUser = authenticate(username= username, password = pwd)
            login(request, loginUser)
            return redirect('inicio')
        else:
            context['reg_form'] = form
    return render(request, 'usuarios/registro.html', context)



@login_required
def perfil(request: HttpRequest):
    context['page_title'] = 'Pelfil del Usuario'
    return render(request, 'usuarios/perfil.html', context)


@login_required
def actualizarPerfil(request: HttpRequest):
    context['page_title'] = 'Actualizar Perfil'
    user = User.objects.get(id=request.user.id)
    if not request.method == 'POST':
        form = ActualizarPerfil(instance=user)
        context['form'] = form
        print(form)
    else:
        form = ActualizarPerfil(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil actualizado correctamente")
            return redirect("perfil")
        else:
            context['form'] = form

    return render(request, 'usuarios/gestion_perfil.html', context)


@login_required
def actualizarContrasena(request: HttpRequest):
    context['page_title'] = "Actualizar Contraseña"
    if request.method == 'POST':
        form = ActualizarContrasena(user = request.user, data= request.POST)
        if form.is_valid():
            form.save()
            messages.success(request,"Contraseña actualizada correctamente")
            update_session_auth_hash(request, form.user)
            return redirect("perfil")
        else:
            context['form'] = form
    else:
        form = ActualizarContrasena(request.POST)
        context['form'] = form
    return render(request, 'usuarios/actualizar_contrasena.html', context)

def is_valid_queryparam(param):
    return param != '' and param is not None

@login_required
def sede(request: HttpRequest):
    semaforo = {
        '1': 'bg-primary',
        '2': 'bg-secondary',
        '3': 'bg-success',
    }
    context['page_title'] = "Sedes"
    sedes = Sede.objects.all()
    context['sedes'] = sedes
    context['semaforo'] = semaforo
    return render(request, 'gestion/sede.html', context)



@login_required
def guardar_sede(request: HttpRequest):
    resp = {'status': 'failed', 'msg': ''}
    if request.method == 'POST':
        if (request.POST['id']).isnumeric():
            sede = Sede.objects.get(pk=request.POST['id'])
        else:
            sede = None
        if sede is None:
            form = GuardarSede(request.POST)
        else:
            form = GuardarSede(request.POST, instance=sede)
        if form.is_valid():
            form.save()
            messages.success(request, 'La Sede se ha guardado exitosamente.')
            resp['status'] = 'success'
        else:
            for fields in form:
                for error in fields.errors:
                    resp['msg'] += str(error + "<br>")
    else:
        resp['msg'] = 'No se han guardado datos.'
    return HttpResponse(json.dumps(resp), content_type='application/json')


@login_required
def adm_sede(request: HttpRequest, pk: str | None = None):
    context['page_title'] = "Gestion de Sedes"
    if pk != None:
        sede = Sede.objects.get(id=pk)
        context['sede'] = sede
    else:
        context['sede'] = {}

    return render(request, 'gestion/adm_sede.html', context)



@login_required
def eliminar_sede(request: HttpRequest):
    resp = {'status': 'failed', 'msg': ''}

    if request.method == 'POST':
        try:
            sede = Sede.objects.get(id=request.POST['id'])
            sede.delete()
            messages.success(request, 'La Sede se ha eliminado exitosamente')
            resp['status'] = 'success'
        except Sede.DoesNotExist as err:
            resp['msg'] = 'La Sede no se pudo eliminar'
            print(err)

    else:
        resp['msg'] = 'La Sede no se pudo eliminar'

    return HttpResponse(json.dumps(resp), content_type="application/json")

def sede_excel(request):
    qs = Sede.objects.order_by('sede')
    username = getpass.getuser()
    current_date = datetime.now().strftime("%Y-%m-%d")
    filename = f'Listado Sedes - {username} - {current_date}.xlsx'
    workbook = Workbook()
    workbook.save(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename="' + filename + '"'


    worksheet = workbook.active

    worksheet.merge_cells('A1:C1')

    first_cell = worksheet['A1']

    first_cell.value = "Listado Sedes"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Listado Sedes'

    # Titulos Columnas
    columns = ['ID', 'Nombre', 'Tipo de Sede',]
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        seventh_cell = worksheet['C3']
        seventh_cell.alignment = Alignment(horizontal="right")

    for sedes in qs:
        row_num += 1

        # Define the data for each cell in the row

        row = [sedes.id, sedes.sede, sedes.tipo,]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook

    workbook.save(response)
    return response


# bus
@login_required
def vehiculo(request: HttpRequest):
    context['page_title'] = "Vehículos"
    vehiculos = Vehiculo.objects.all()
    context['vehiculos'] = vehiculos

    return render(request, 'gestion/vehiculo.html', context)

@login_required
def guardar_vehiculo(request:HttpRequest):
    resp = {'status': 'failed', 'msg': ''}
    if request.method == 'POST':
        if (request.POST['id']).isnumeric():
            vehiculo = Vehiculo.objects.get(pk=request.POST['id'])
        else:
            vehiculo = None
        if vehiculo is None:
                form = GuardarVehiculo(request.POST)
        else:
            form = GuardarVehiculo(request.POST, instance=vehiculo)
        if form.is_valid():
            form.save()
            messages.success(request, 'El vehículo se ha guardado exitosamente')
            resp['status'] = 'success'
        else:
            for fields in form:
                for error in fields.errors:
                    resp['msg'] += str(error + "<br>")
    else:
        resp['msg'] = 'No se han guardado datos.'
    return HttpResponse(json.dumps(resp), content_type='application/json')


@login_required
def adm_vehiculo(request: HttpRequest, pk: str | None = None):
    context['page_title'] = "Gestión de Buses"
    propietarios = Propietario.objects.all()
    context['propietarios'] = propietarios

    if pk != None:
        vehiculo = Vehiculo.objects.get(id=pk)
        context['vehiculo'] = vehiculo
    else:
        context['vehiculo'] = {}

    return render(request, 'gestion/adm_vehiculo.html', context)


@login_required
def eliminar_vehiculo(request: HttpRequest):
    resp = {'status': 'failed', 'msg': ''}
    #print(f"datos: {request}")
    if request.method == 'POST':

        try:
            vehiculo = Vehiculo.objects.get(id=request.POST['id'])
            vehiculo.delete()
            messages.success(request, 'El vehículo se ha eliminado exitosamente.')
            resp['status'] = 'success'
        except Vehiculo.DoesNotExist as err:
            resp['msg'] = 'El Vehículo no se pudo eliminar'
            print(err)
    else:
        resp['msg'] = 'El Vehículo no se pudo eliminar'
    return HttpResponse(json.dumps(resp), content_type="application/json")



def vehiculo_excel(request):
    qs = Vehiculo.objects.order_by('numero_veh')
    username = getpass.getuser()
    current_date = datetime.now().strftime("%Y-%m-%d")
    filename = f'Listado Vehículos - {username} - {current_date}.xlsx'
    workbook = Workbook()
    workbook.save(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename="' + filename + '"'


    worksheet = workbook.active

    worksheet.merge_cells('A1:G1')

    first_cell = worksheet['A1']

    first_cell.value = "Listado de vehículos"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Listado Vehículos'

    # Titulos Columnas
    columns = ['ID', 'Fecha de Ingreso', 'Propietario', 'Placa', 'Número de Vehiculo', 'Asientos', 'Estado',]
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        seventh_cell = worksheet['G3']
        seventh_cell.alignment = Alignment(horizontal="right")

    for vehiculos in qs:
        row_num += 1

        # Define the data for each cell in the row


        row = [vehiculos.id,vehiculos.fecha_creado.strftime("%Y%m%d - %H:%M:%S"),vehiculos.propietario.propietario,vehiculos.placa_veh,vehiculos.numero_veh,vehiculos.asientos,vehiculos.estado,]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook

    workbook.save(response)
    return response



@login_required
def programacion(request: HttpRequest):
    semaforo_2 = {
        '0': 'bg-danger',
        '1': 'bg-dark',
        '2': 'bg-warning',
    }

    context['page_title'] = "Progamación"
    programaciones = Programacion.objects.all()
    context['programaciones'] = programaciones
    context['semaforo_2'] = semaforo_2
    return render(request,'gestion/programacion.html',context)



@login_required
def guardar_programacion(request: HttpRequest):
    resp = {'status':'failed','msg':''}
    if request.method == 'POST':
        if (request.POST['id']).isnumeric():
            programacion = Programacion.objects.get(pk=request.POST['id'])
        else:
            programacion = None
        if programacion is None:
                form = GuardarProgramacion(request.POST)
        else:
            form = GuardarProgramacion(request.POST,instance=programacion)
        if form.is_valid():
            form.save()
            messages.success(request, 'La Programación se ha guardado exitosamente.')
            resp['status'] = 'success'
        else:
            print(form.errors)
            for fields in form:
                for error in fields.errors:
                    resp['msg'] += str(error + "<br>")
    else:
        resp['msg'] = 'No se han guardado datos.'
    return HttpResponse(json.dumps(resp), content_type='application/json')



@login_required
def adm_programacion(request: HttpRequest, pk: str | None = None):
    context['page_title'] = "Gestión Programación"
    sedes = Sede.objects.all()
    vehiculos = Vehiculo.objects.filter(estado=1).all()
    conductores = Conductor.objects.filter(estado=1).all()
    context['vehiculos'] = vehiculos
    context['sedes'] = sedes
    context['conductores'] = conductores

    if pk != None:
        programacion = Programacion.objects.get(id=pk)
        context['programacion'] = programacion
    else:
        context[' programacion'] = {}

    return render(request, 'gestion/adm_programacion.html', context)



@login_required
def eliminar_programacion(request: HttpRequest):
    resp = {'status': 'failed', 'msg': ''}
    if request.method == 'POST':
        try:
            programacion = Programacion.objects.get(id=request.POST['id'])
            programacion.delete()
            messages.success(request, 'La Programación ha eliminado exitosamente')
            resp['status'] = 'success'
        except Programacion.DoesNotExist as err:
            resp['msg'] = 'La Programación no se pudo eliminar'
            print(err)

    else:
        resp['msg'] = 'La Programación no se pudo eliminar'

    return HttpResponse(json.dumps(resp), content_type="application/json")



def programacion_excel(request):
    qs = Programacion.objects.order_by('programacion')
    username = getpass.getuser()
    current_date = datetime.now().strftime("%Y-%m-%d")
    filename = f'Listado Programaciones - {username} - {current_date}.xlsx'
    workbook = Workbook()
    workbook.save(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename="' + filename + '"'


    worksheet = workbook.active

    worksheet.merge_cells('A1:I1')

    first_cell = worksheet['A1']

    first_cell.value = "Listado Programaciones"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Listado Programaciones'

    # Titulos Columnas
    #  Llamar al ultimo usuario en editar (Coordina)
    columns = ['ID','Coordina','Conductor','Horario', 'Código', 'Estado', 'Origen', 'Destino', 'Precio', ]
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        seventh_cell = worksheet['J3']
        seventh_cell.alignment = Alignment(horizontal="right")

    for programaciones in qs:
        row_num += 1


        # Define the data for each cell in the row

        # username el ultimo que asigna algun cambio
        row = [programaciones.id, '{user.username}',programaciones.conductor.conductor, programaciones.programacion.strftime("%Y%m%d - %H:%M:%S"), programaciones.codigo, programaciones.estado,
               programaciones.origen.sede, programaciones.destino.sede, programaciones.precio,]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook

    workbook.save(response)
    return response



@login_required
def encomienda(request: HttpRequest):
    semaforo_3 = {
        '1': 'bg-danger',
        '2': 'bg-dark',
        '3': 'bg-warning',
        '4': 'bg-succes',

    }

    context['page_title'] = "Encomienda"
    encomiendas = Encomienda.objects.all()
    context['encomiendas'] = encomiendas
    context['semaforo_3'] = semaforo_3
    return render(request, 'gestion/encomienda.html', context)


@login_required
def guardar_encomienda(request):
     resp = {'status':'failed','msg':''}
     if request.method == "POST":
         form= GuardarEncomienda(request.POST)
         if form.is_valid():
           form.save()
           messages.success(request, 'La Encomienda se ha guardado exitosamente.')
           resp['status'] = 'success'

         else:
            print("Error")
     else:

        form= GuardarProgramacion()
        context ={
            'form':form
        }

     return HttpResponse(json.dumps(resp), content_type='application/json')



@login_required
def adm_encomienda(request: HttpRequest, pk: str | None = None):
    context['page_title'] = "Gestión de Encomiendas"
    programaciones =Programacion.objects.all()
    context['programaciones'] = programaciones
    if pk != None:
        encomienda = Encomienda.objects.get(id=pk)
        context['encomienda'] = encomienda
    else:
        context['encomienda'] = {}

    return render(request, 'gestion/adm_encomienda.html', context)



@login_required
def eliminar_encomienda(request: HttpRequest):
    resp = {'status': 'failed', 'msg': ''}
    if request.method == 'POST':
        try:
            encomienda = Encomienda.objects.get(id=request.POST['id'])
            encomienda.delete()
            messages.success(request, 'Encomienda eliminada exitosamente.')
            resp['status'] = 'success'
        except Encomienda.DoesNotExist as err:
            resp['msg'] = 'Encomienda no se pudo eliminar'
            print(err)
    else:
        resp['msg'] = 'Encomienda no se pudo eliminar'
    return HttpResponse(json.dumps(resp), content_type="application/json")

def encomienda_excel(request):

    qs = Encomienda.objects.order_by('id')
    username = getpass.getuser()
    current_date = datetime.now().strftime("%Y-%m-%d")
    filename = f'Listado Encomiendas - {username} - {current_date}.xlsx'
    workbook = Workbook()
    workbook.save(filename)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
    response['Content-Disposition'] = 'attachment; filename="' + filename + '"'


    worksheet = workbook.active

    worksheet.merge_cells('A1:H1')


    first_cell = worksheet['A1']

    first_cell.value = "Listado Encomienda"
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Listado Encomienda'
    #  ultimo en editar
    # Titulos Columnas
    columns = ['ID','Facilitador','Programacion', 'Nombre', 'Telefono','Cedula', 'Nombre', 'Apellido', 'Telefono', 'Costo de Envio', 'Estado', 'Ultima Actualización',]
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font = Font(bold=True, color="F7F6FA")
        seventh_cell = worksheet['h3']
        seventh_cell.alignment = Alignment(horizontal="right")

    for encomiendas in qs:
        row_num += 1

        # Define the data for each cell in the row
        # username el ultimo que asigna algun cambio

        row = [encomiendas.id,'{user.username}','{encomiendas.programacion}', encomiendas.nombre_envio,encomiendas.telefono_envio,encomiendas.cedula_envio,
        encomiendas.nombre_recibido,encomiendas.telefono_recibido,encomiendas.cedula_recibido,encomiendas.costo_envio,
        encomiendas.estado, encomiendas.fecha_actualizado.strftime("%Y%m%d - %H:%M:%S"), ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook

    workbook.save(response)
    return response


def buscar_programado(request: HttpRequest):
    context['page_title'] = 'Buscar Viajes Programacions'
    context['sedes'] = Sede.objects.filter(tipo__gt = 0).all
    today = datetime.today().strftime("%Y-%m-%d")
    context['today'] = today
    return render(request, 'busqueda/buscar_viaje.html', context)

def viajes_programados(request: HttpRequest):
    if not request.method == 'POST':
        context['page_title'] = "Viajes Programacions"
        programaciones = Programacion.objects.filter(estado=1, programacion__gt=datetime.now()).all()
        context['programaciones'] = programaciones
        context['is_searched'] = False
        context['data'] = {}
    else:
        context['page_title'] = "Resultados | Viajes Programacions"
        context['is_searched'] = True
        date = datetime.strptime(request.POST['date'],"%Y-%m-%d").date()
        year= date.strftime('%Y')
        month = date.strftime('%m')
        day = date.strftime('%d')
        origen = Sede.objects.get(id=request.POST['origen'])
        destino = Sede.objects.get(id=request.POST['destino'])
        programaciones = Programacion.objects.filter(Q(estado=1) & Q(programacion__year=year) & Q(programacion__month=month) & Q(programacion__day=day) & Q(Q(origen=origen) | Q(destino=destino))).all()
        context['programaciones'] = programaciones
        context['data'] = {'date':date,'origen': origen, 'destino': destino}

    return render(request, 'busqueda/viajes_programados.html', context)

def base(request: HttpRequest):
    call_command('dbbackup')
    return HttpResponse('db creada')


def contacto(request: HttpRequest):
    context={}
    return render(request, 'core/contacto.html', context)


def error_404(request, exception):
    data = {"name": "ThePythonDjango.com"}
    return render(request,'core/error_404.html', data)
